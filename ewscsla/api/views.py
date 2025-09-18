import django_filters
from django.db.models import Count
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import viewsets, status
from rest_framework.decorators import permission_classes, api_view
from rest_framework.permissions import IsAuthenticated
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework import generics
from django.contrib.auth import logout
from rest_framework.views import APIView

from api.pagination import ListPagination
from api.serializers import SlaEntrySerializer, DepartmentSerializer, ListSlaEntrySerializer, ListSlaRatingSerializer, \
    SlaRatingSerializer, SlaImprovementPlanEntrySerializer, SlaCustomerStatusEntrySerializer, AuthUserSerializer, \
    AggregatedRatingsSerializer, AggregatedDepartmentDataSerializer
from core.models import SlaEntry, Department, SlaRating, SlaImprovementPlanEntry, SlaCustomerStatusEntry, AuthUser


class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.order_by('pk')
    permission_classes = [IsAuthenticated, ]
    serializer_class = DepartmentSerializer


# ======================================================================

class SlaEntryViewSet(viewsets.ModelViewSet):
    queryset = SlaEntry.objects.order_by('pk')
    permission_classes = [IsAuthenticated, ]
    filterset_fields = ['department', ]

    def get_serializer_class(self):
        if self.action == 'list' or self.action == 'retrieve':
            return ListSlaEntrySerializer
        return SlaEntrySerializer


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_sla_entry(request: Request):
    data = {
        **request.data,
        'added_by': request.user.pk
    }
    form = SlaEntrySerializer(data=data, many=False)
    if form.is_valid():
        entry = form.save()
        return Response(
            data=SlaEntrySerializer(instance=entry).data,
            status=status.HTTP_200_OK
        )

    return Response(
        data=form.errors,
        status=status.HTTP_400_BAD_REQUEST
    )


# ======================================================================

class SlaRatingFilter(django_filters.FilterSet):
    department = django_filters.CharFilter(
        field_name='sla__department', lookup_expr='exact')

    class Meta:
        model = SlaRating
        fields = ('department',)


class UserSlaRatingEntriesView(generics.ListAPIView):
    permission_classes = [IsAuthenticated, ]
    filterset_class = SlaRatingFilter
    serializer_class = ListSlaRatingSerializer

    def get_queryset(self):
        return SlaRating.objects.filter(
            rated_by=self.request.user
        ).order_by('pk')


class SlaRatingEntryViewSet(viewsets.ModelViewSet):
    queryset = SlaRating.objects.order_by('pk')
    permission_classes = [IsAuthenticated, ]
    pagination_class = ListPagination
    filterset_class = SlaRatingFilter

    def get_queryset(self):
        queryset = self.queryset
        filter_by = self.request.query_params.get('filter_by')
        if filter_by == 'my_entries':
            # Filter by the current user's SLA entries
            queryset = queryset.filter(sla__added_by=self.request.user)
        return queryset.order_by('pk')

    def get_serializer_class(self):
        if self.action == 'list' or self.action == 'retrieve':
            return ListSlaRatingSerializer
        return SlaRatingSerializer


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_sla_rating_entry(request: Request):
    data = {
        **request.data,
        'rated_by': request.user.pk
    }
    form = SlaRatingSerializer(data=data, many=False)
    if form.is_valid():
        rating = form.save()
        return Response(
            data=SlaRatingSerializer(instance=rating).data,
            status=status.HTTP_200_OK
        )

    return Response(
        data=form.errors,
        status=status.HTTP_400_BAD_REQUEST
    )


# ======================================================================

class SlaImprovementActionPlanViewSet(viewsets.ModelViewSet):
    queryset = SlaImprovementPlanEntry.objects.order_by('pk')
    permission_classes = [IsAuthenticated, ]
    serializer_class = SlaImprovementPlanEntrySerializer


# ======================================================================

class SlaCustomerStatusPlanViewSet(viewsets.ModelViewSet):
    queryset = SlaCustomerStatusEntry.objects.order_by('pk')
    permission_classes = [IsAuthenticated, ]
    serializer_class = SlaCustomerStatusEntrySerializer


@api_view(['GET'])
def generate_report(request):
    try:
        today = timezone.now().strftime("%Y_%m_%d_%H_%M")

        workbook = Workbook()
        worksheet = workbook.active

        # Set page properties
        worksheet.title = "SLA Report"
        worksheet.page_setup.fitToPage = True
        worksheet.page_setup.fitToHeight = 0
        worksheet.page_setup.fitToWidth = 1

        # Define styles
        title_font = Font(name='Calibri', size=16, bold=True)
        header_font = Font(
            name='Calibri', size=11, bold=True,
            color="FFFFFF")  # White font color
        header_fill = PatternFill(
            start_color="3399FF", end_color="3399FF", fill_type="solid")
        header_alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)
        data_alignment = Alignment(vertical='center', wrap_text=True)

        # worksheet.append(["ESWATINI WATER SERVICES CORPORATION"])
        # worksheet.append([])
        # worksheet.append(["SLA's MONITORING TOOL"])

        # Add report titles with merged cells and styling
        worksheet.merge_cells('A1:M1')
        title_cell_1 = worksheet['A1']
        title_cell_1.value = "ESWATINI WATER SERVICES CORPORATION"
        title_cell_1.font = title_font
        title_cell_1.alignment = Alignment(horizontal='center')

        worksheet.merge_cells('A3:M3')
        title_cell_2 = worksheet['A3']
        title_cell_2.value = "SLA's MONITORING TOOL"
        title_cell_2.font = title_font
        title_cell_2.alignment = Alignment(horizontal='center')

        # Add an empty row for spacing
        worksheet.append([])

        # worksheet.append([
        #     "Internal Service Provider", "Internal Customer", "Report Qrt Date",
        #     "Rating", "SLA", "Reason for rating", "Agreed Improvement Action", "Due Date",
        #     "Manager Status (Service Provider)", "Internal Customer Status"
        # ])

        main_headers = [
            "Internal Service Provider",
            "Internal Customer",
            "Report Qrt Date",
            "Service Provider Responsibility",
            "Customer Responsibility",
            "Service Level",
            "Rating",
            "SLA",
            "Reason for rating",
            "Agreed Improvement Action",
            "Due Date",
            "Manager Status (Service Provider)",
            "Internal Customer Status"
        ]
        worksheet.append(main_headers)

        # Apply styles to the header row
        for col_num, _ in enumerate(main_headers, 1):
            col_letter = get_column_letter(col_num)
            cell = worksheet[f"{col_letter}5"]
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill

        # Set a fixed width for the columns with long text paragraphs
        # Corresponds to SP Responsibility, Customer Responsibility, Service Level, Reason for rating, Agreed Improvement Action
        long_text_columns = [4, 5, 6, 8, 9, 10]
        fixed_width = 30
        for col_num in long_text_columns:
            col_letter = get_column_letter(col_num)
            worksheet.column_dimensions[col_letter].width = fixed_width

        # Add data rows
        sla_ratings = SlaRating.objects.order_by('updated_at')
        for rating in sla_ratings:
            sla_entry = rating.sla

            # Data for each cell
            service_provider = sla_entry.department.name
            customer = rating.rated_by.department.name \
                if (rating.rated_by and rating.rated_by.department) else "N/A"
            sla_date = sla_entry.date.strftime("%Y-%m-%d")
            sp_responsibility = sla_entry.service_provider_responsibility or "N/A"
            cust_responsibility = sla_entry.customer_responsibility or "N/A"
            service_level = sla_entry.service_level or "N/A"
            rating_value = str(int(rating.rating))
            key_performance_area = sla_entry.key_performance_area or "N/A"
            reason = rating.reason or "N/A"
            action_plan = rating.action_plan.improvement_action \
                if hasattr(rating, 'action_plan') else "N/A"
            due_date = rating.action_plan.due_date.strftime(
                "%Y-%m-%d") if hasattr(rating, 'action_plan') else "N/A"
            manager_status = rating.action_plan.get_status_display() \
                if hasattr(rating, 'action_plan') else "N/A"
            customer_status = rating.customer_status.get_status_display() \
                if hasattr(rating, 'customer_status') else "N/A"

            row_data = [
                service_provider, customer, sla_date, sp_responsibility, cust_responsibility,
                service_level, rating_value, key_performance_area, reason,
                action_plan, due_date, manager_status, customer_status
            ]

            worksheet.append(row_data)

            # Apply data alignment for the current row, especially for text wrapping
            for col_num in range(1, len(row_data) + 1):
                col_letter = get_column_letter(col_num)
                worksheet[f"{col_letter}{worksheet.max_row}"].alignment = data_alignment

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=SLA_REPORT_{today}.xlsx'
        response.write(save_virtual_workbook(workbook))

        return response

        # sla_ratings = SlaRating.objects.order_by('updated_at')
        # for rating in sla_ratings:
        #     service_provider = rating.sla.department.name
        #     customer = rating.rated_by.department.name if (
        #         rating.rated_by and rating.rated_by.department) else "N/A"
        #     sla_date = rating.sla.date.strftime("%Y-%m-%d")
        #     rating_value = str(int(rating.rating))
        #     key_performance_area = rating.sla.key_performance_area
        #     reason = rating.reason
        #     action_plan = rating.action_plan.improvement_action if hasattr(
        #         rating, 'action_plan') else "N/A"
        #     due_date = rating.action_plan.due_date.strftime(
        #         "%Y-%m-%d") if hasattr(rating, 'action_plan') else "N/A"
        #     manager_status = rating.action_plan.get_status_display() if hasattr(rating,
        #                                                                         'action_plan') else "N/A"
        #     customer_status = rating.customer_status.get_status_display() \
        #         if hasattr(rating, 'customer_status') else "N/A"

        #     worksheet.append([
        #         service_provider, customer, sla_date, rating_value, key_performance_area, reason,
        #         action_plan, due_date, manager_status, customer_status
        #     ])

        # response = HttpResponse(
        #     content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # response['Content-Disposition'] = f'attachment; filename=SLA_REPORT_{today}.xlsx'
        # response.write(save_virtual_workbook(workbook))

        # return response

    except Exception as error:
        return Response({
            'error': 'Failed to generate report',
            'debug': str(error)
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
def export_sla_entries(request):
    try:
        today = timezone.now().strftime("%Y_%m_%d_%H_%M")

        workbook = Workbook()
        worksheet = workbook.active

        worksheet.append(["ESWATINI WATER SERVICES CORPORATION"])
        worksheet.append([])
        worksheet.append(["SLA's MONITORING TOOL - SLA ENTRIES"])

        worksheet.append([
            "SLA Department", "Service Description", "Customer Responsibility",
            "Service Level", "SLA Date", "Added By",
        ])

        sla_entries = SlaEntry.objects.order_by('department__name')
        for sla in sla_entries:
            worksheet.append([
                sla.department.name,
                sla.key_performance_area,
                sla.customer_responsibility,
                sla.service_level,
                sla.date.strftime("%Y-%m-%d"),
                sla.added_by.get_full_name() or sla.added_by.username
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response[
            'Content-Disposition'] = f'attachment; filename=SLA_ENTRIES_REPORT_{today}.xlsx'
        response.write(save_virtual_workbook(workbook))

        return response

    except Exception as error:
        return Response({
            'error': 'Failed to generate report',
            'debug': str(error)
        }, status=status.HTTP_400_BAD_REQUEST)


# ======================================================================

@api_view(['GET'])
def dashboard(request):
    try:
        users = AuthUser.objects.count()
        sla_entries = SlaEntry.objects.count()
        ratings = SlaRating.objects.count()
        action_plans = SlaImprovementPlanEntry.objects.count()

        departments = Department.objects.order_by("pk")
        departments_data = departments.annotate(
            employees_count=Count('employees', distinct=True),
            sla_entries_count=Count('sla_entries', distinct=True)
        )

        aggregated_ratings = []
        for department in departments:
            department_ratings = SlaRating.objects.filter(
                sla__department=department
            ).values('rating').annotate(rating_count=Count("rating"))

            _grouped_ratings = {}
            for _rating in department_ratings:
                if _rating["rating"] == 1:
                    _grouped_ratings.update({
                        "Met_None": _rating["rating_count"],
                    })
                elif _rating["rating"] == 2:
                    _grouped_ratings.update({
                        "Met_Some": _rating["rating_count"],
                    })
                elif _rating["rating"] == 3:
                    _grouped_ratings.update({
                        "Met_All": _rating["rating_count"],
                    })
                elif _rating["rating"] == 4:
                    _grouped_ratings.update({
                        "Exceeded_Some": _rating["rating_count"],
                    })
                elif _rating["rating"] == 5:
                    _grouped_ratings.update({
                        "Exceeded_All": _rating["rating_count"],
                    })

            aggregated_ratings.append({
                'department': department.name,
                "ratings": _grouped_ratings
            })

        return Response({
            'users': users,
            'sla_entries': sla_entries,
            'ratings': ratings,
            'action_plans': action_plans,
            'aggregated_ratings': AggregatedRatingsSerializer(aggregated_ratings, many=True).data,
            'department_data': AggregatedDepartmentDataSerializer(departments_data, many=True).data
        }, status=status.HTTP_200_OK)

    except Exception as error:
        return Response({
            'error': str(error)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def profile(request):
    try:
        user = AuthUserSerializer(instance=request.user)
        return Response(data=user.data, status=status.HTTP_200_OK)

    except Exception as error:
        return Response({
            'error': str(error)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class Logout(APIView):
    def get(self, request):
        logout(request)
        return Response(status=status.HTTP_200_OK)
